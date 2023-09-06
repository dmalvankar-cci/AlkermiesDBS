from openpyxl.reader.excel import load_workbook

def read_data(rowNum, colNum):
        wrkBk = load_workbook("Alkermies_creds.xlsx")
        sheet = wrkBk.get_sheet_by_name("Sheet1")
        return sheet.cell(row=rowNum, column=colNum).value

